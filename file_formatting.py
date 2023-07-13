from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from openpyxl import Workbook
from time import strftime, gmtime


class FileFormatting:

    def __init__(self, file_name):
        """
        Initialize the FileFormatting class.

        Args:
            file_name (str): The name of the Excel file.
        """
        self.file_name = file_name
        self.bibliotheque = {
            "A": 1,
            "B": 2,
            "C": 3,
            "D": 4,
            "E": 5,
            "F": 6,
            "G": 7,
            "H": 8,
            "I": 9,
            "J": 10
        }
        self.feuille = None
        self.classeur = None

    def load_file(self):
        """
        Create a new Excel workbook and select the first sheet.

        """
        self.classeur = Workbook()
        self.feuille = self.classeur.active
        
    def save_file(self):
        """
        Save the workbook to a file.

        """
        self.classeur.save(self.file_name)
        print(f"The file {self.file_name} has been saved successfully.")

    def value_column(self, choice):
        """
        Get the numerical value of a column based on its letter representation.

        Args:
            choice (str): The letter representation of the column.

        Returns:
            int: The numerical value of the column.

        """
        valeur = self.bibliotheque[choice]
        return valeur

    def width_adjustment(self, lettre, size):
        """
        Adjust the width of a column.

        Args:
            lettre (str): The letter representation of the column.
            size (int): The desired width of the column.

        """
        colonne_ = self.feuille.column_dimensions[get_column_letter(self.value_column(lettre))]
        colonne_.width = size

    def headers(self, list, ligne):
        """
        Add headers to a row in the worksheet.

        Args:
            list (list): The list of header values.
            ligne (int): The row number where the headers should be added.

        """
        for col, valeur in enumerate(list, 1):
            self.feuille.cell(row=ligne, column=col, value=valeur)
        for cell in self.feuille[ligne]:
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)

    def add_line(self, line, line_number):
        """
        Add a line of values to the worksheet.

        Args:
            line (list): The list of values to be added as a line.
            line_number (int): The row number where the line should be added.

        """
        for col, val in enumerate(line, start=1):
            self.feuille.cell(row=line_number, column=col, value=val)

    def alignment_line(self, line_number):
        """
        Apply horizontal center alignment to all cells in a row.

        Args:
            line_number (int): The row number of the line to be aligned.

        """
        for cell in self.feuille[line_number]:
            cell.alignment = Alignment(horizontal='center')

    def add_image(self, path_image, cell):
        """
        Add an image to a cell in the worksheet.

        Args:
            path_image (str): The path to the image file.
            cell (str): The cell reference where the image should be added.

        """
        img = Image(path_image)
        self.feuille.add_image(img, cell)

    def merge_cells(self, start_line, number_of_lines):
        """
        Merge cells in a range of rows.

        Args:
            start_line (int): The starting row number of the cell range.
            number_of_lines (int): The number of rows to be merged.

        """
        for i in range(start_line, start_line + number_of_lines):
            self.feuille.merge_cells(f'M{i}:u{i}')


    def Resume_info(self, total_Frames, Fps, frame_firstdetect, crop, radius):
        """
        Add related information of video to the worksheet.

        Args:
            total_Frames (int): The total number of frames in the video.
            Fps (int): Frames per second of the video.
            frame_firstdetect (int): The frame number of the first detection.
            crop (bool): Crop information.
            radius (int): The radius of the circle.

        """
        self.feuille['M3'] = "Info Video"
        cell = self.feuille['M3']
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

        self.feuille['M4'] = f"ID Video: {str(input('ID de la video: '))}"
        time_duration = total_Frames / Fps
        self.feuille['M5'] = f"Durée de video : {strftime('%H:%M:%S', gmtime(time_duration))}"
        self.feuille['M6'] = f"Frame de la premiére detection : {str(frame_firstdetect)}"
        time_duration = frame_firstdetect / Fps
        self.feuille['M7'] = f"timer de la premiére detection : {strftime('%H:%M:%S', gmtime(time_duration))}"
        self.feuille['M8'] = f"Rayon du cercle : {radius}"
        self.feuille['M9'] = "bouteille explosion : "
        self.feuille['M10'] = "duré de la  detection : "
        self.feuille['M11'] = f"Crop: {crop}"